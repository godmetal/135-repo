import requests
import datetime
from datetime import timedelta
from parse import *
import re

# SSL Warning 메세지 Off
import urllib3

urllib3.disable_warnings()

try:
    from collections.abc import Mapping
except ImportError:
    from collections import Mapping

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

import datetime
from datetime import timedelta

import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Color, NamedStyle
from openpyxl.chart import BarChart, Reference
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

from string import ascii_uppercase

# 테두리 스타일
paintstyle = NamedStyle(name="border")
bd = Side(border_style='thin', color="000000")
paintstyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)
paintstyle.alignment = Alignment(vertical='center')


# 시트 첫행 타이틀 굵게, 가운데정렬, 배경색설정
def setTitleCell(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(patternType='solid', fgColor=Color("FFFF00"))
    return cell


# 테이블 제목 (볼드, 회색배경)
def writeColName(cName, row, ws):
    column_char = 'a'
    # write column name
    for name in cName:
        ws[column_char + str(row)] = name
        ws[column_char + str(row)].style = paintstyle
        ws[column_char + str(row)].fill = PatternFill(patternType='solid', fgColor=Color("e1e5eb"))
        ws[column_char + str(row)].font = Font(name='맑은 고딕', size=10, bold=True)
        # ws[column_char + str(row)].alignment = Alignment(horizontal='center', vertical='center')
        column_char = chr(ord(column_char) + 1)


# 본문 필드
def saveContent(cName, ws):
    column_char = 'a'
    for name in cName:
        ws[column_char + str(row_num)] = name
        ws[column_char + str(row_num)].style = paintstyle
        ws[column_char + str(row_num)].alignment = Alignment(vertical='center', wrap_text=True)
        ws[column_char + str(row_num)].font = Font(name='맑은 고딕', size=10, bold=False)
        column_char = chr(ord(column_char) + 1)


# 함수 : 쌍따옴표, EN DASH 제거
def cleanText(readData):
    text = re.sub('\"', '', readData)
    # text = re.sub('–', '-', text)
    # text = re.sub('•', '-', text)
    return text


# UTF-8을 EUC-KR로 변환
def utf2euc(str):
    return unicode(str, 'utf-8').encode('euc-kr')


# EUC-KR을 UTF-8로 변환
def euc2utf(str):
    return unicode(str, 'euc-kr').encode('utf-8')


def ensureUtf(s):
    try:
        if type(s) == unicode:
            return s.encode('utf8', 'ignore')
    except:
        return str(s)


#################################################################
### SPLUNK 연동


# 세션 KEY 받아오기
sessionURL = "https://52.2.138.27:8089/services/auth/login"

# defining a params dict for the parameters to be sent to the API
PARAMS = {'username': 'mailscreen', 'password': 'secu12!@'}
# PARAMS2 = {'username':'mailscreen','password':'secu12!@'}

# sending get request and saving the response as response object
r = requests.post(url=sessionURL, data=PARAMS, verify=False, json='xml')

sessionKey = parse("<response>\n  <sessionKey>{}</sessionKey>\n</response>", r.text)
r.connection.close()

# 검색 조건 정의
past = datetime.datetime.now() - timedelta(days=7)  # 검색시작
earliest_time = past.strftime('%Y-%m-%d') + 'T' + past.strftime('%H:%M:%S')

now = datetime.datetime.now()  # 검색종료
latest_time = now.strftime('%Y-%m-%d') + 'T' + now.strftime('%H:%M:%S')

# 로그 조회 URL
logURL = "https://52.2.138.27:8089/services/search/jobs/export"

# 세션KEY 를 Herder 정보에 추가
headerInfo = {'Authorization': 'Splunk ' + sessionKey[0]}
#


# 메일 발송
now = datetime.datetime.now()
today = now.strftime('%Y-%m-%d')
yesterday = datetime.datetime.now() - timedelta(days=1)
yesterday = yesterday.strftime('%Y-%m-%d')
oneweekday = datetime.datetime.now() - timedelta(days=7)
oneweekday = oneweekday.strftime('%Y-%m-%d')
twoweekday = datetime.datetime.now() - timedelta(days=14)
twoweekday = twoweekday.strftime('%Y-%m-%d')

#################################################################
# 본문 상단 및 메일 스타일 선언, 제목
HTML = """
<html>
<head>
<style id="NamoSE__ParagraphTagStyle" type="text/css">p:not(.ce_exstyle){font-size:10pt;font-family:바탕;line-height:1.2;} td{font-family:''; font-size:10pt;} h1,h2,h3,h4,h5,h6{font-family:'';}
</style>
<meta http-equiv="Content-Type" content="text/html; charset=utf-8" />

<style id="NamoSE__GeneralStyle" type="text/css" data-namo-bodystyle="data-namo-bodystyle"> body{font-family :굴림; color : #000000; font-size : 10pt; margin: 7px 0 0 7px;} p,li{line-height:1.2; word-wrap: break-word; margin-top:0; margin-bottom:0;} body{overflow:auto;}.NamoSE_layoutlock_show { word-break: break-all;} 
</style>
</head>
<body>
<p><strong><u><span style="font-family: 바탕; font-size: 14pt; background-color: rgb(255, 242, 204);"><br /></span></u></strong></p>
<p class="ce_exstyle" style="text-align: left; line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong><u><span style="font-family: 바탕; font-size: 14pt; background-color: rgb(255, 242, 204);">■ 메일스크린 주간</span><span style="font-family: 바탕; font-size: 14pt; background-color: rgb(255, 242, 204);"> 리포트 ("""

HTML = HTML + oneweekday + """ ~ """ + yesterday

HTML = HTML + """)</span></u></strong></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-size: 11pt;">&nbsp;</span></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-size: 11pt;">&nbsp;- 금주 기준 : """ + oneweekday + """ 00시 00분 ~ """ + yesterday + """ 23시 59분</p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;">&nbsp;- 메일 집계 수량 </span><span style="font-family: 바탕; font-size: 11pt;">: 발신자(보내는메일) 기준으로 수량이 집계됨.</span></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><br /></p>

"""
#################################################################
# 제목 및 테이블 상단 작성 : 1. 메일스크린 정책 적용 현황
HTML = HTML + """
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong><span style="font-family: 바탕; font-size: 12pt;">1. 메일스크린 정책 적용 현황</span></strong></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><br /></p>
<table style="border: 0px solid rgb(0, 0, 0); border-image: none; width: 609px; height: 154px; font-size: 10pt; border-collapse: collapse; background-color: rgb(255, 255, 255);" border="1" cellspacing="0" cellpadding="0"> 
<tbody> 
<tr> 
<td style="border: 1px solid rgb(0, 0, 0); border-image: none; width: 241px; height: 28px; background-color: rgb(204, 204, 204);"> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong>정책 구분</strong></p> </td> 
<td style="border: 1px solid rgb(0, 0, 0); border-image: none; width: 121px; height: 28px; background-color: rgb(204, 204, 204);"> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong>금주 발송 수량</strong></p> </td> 
<td style="border: 1px solid rgb(0, 0, 0); border-image: none; width: 121px; height: 28px; background-color: rgb(204, 204, 204);"> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong>전주 발송 수량<br /></strong></p> </td> 
<td style="border: 1px solid rgb(0, 0, 0); border-image: none; width: 121px; height: 28px; background-color: rgb(204, 204, 204);"> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong>증감(+/-)</strong></p> </td> 
</tr> """

#################################################################
# 본문작성을 위한 UBA 로그 연동 : 1. 메일스크린 정책 적용 현황
splunksql = """search index=mailscreen_group policyname="올리브영*" earliest="-7d@d" latest="@d"
 |rename sender_email as email_addr
 |fields sendtime, email_addr, header_subject, policyname
 |dedup sendtime, email_addr, header_subject, policyname
 |eventstats count(policyname) as count1 by policyname
 | dedup policyname

 | append

 [ search index=mailscreen_group policyname="올리브영*" earliest="-14d@d" latest="-7d@d"
 |rename sender_email as email_addr
 |fields sendtime, email_addr, header_subject, policyname]
 |dedup sendtime, email_addr, header_subject, policyname
 |eventstats count(policyname) as count2 by policyname
 | dedup policyname

 | table policyname, count1, count2
 | fillnull value=0 count1, count2
 | sort - count1"""

PARAMS = {'search': splunksql,
          'earliest_time': earliest_time,
          'latest_time': latest_time,
          'output_mode': 'json'}
# 응답 요청
r = requests.post(url=logURL, headers=headerInfo, data=PARAMS, verify=False)

with open(now.strftime('%Y-%m-%d') + '_oyscreen.csv', 'wb') as file:
    file.write(r.content)

with open(now.strftime('%Y-%m-%d') + '_oyscreen.csv', "rt", encoding="UTF8") as infile:
    for line in infile:

        # 데이터를 쉼표 로 구분
        data = line.split(',\"')
        # print(line)
        # preview:true는 버림. 왜 조회 결과가 2쌍으로 나오는지 모르곘음...
        # print(cleanText(data[0]))
        if cleanText(data[0]) == "{preview:false":
            # 조회 결과가 1개일 때, {"preview":false,"offset":0,"lastrow":true,"result":{"comp_nm":......
            # print(cleanText(data[2]))
            if cleanText(data[2]) != "lastrow:true":
                # print(line)
                col1 = cleanText(data[2]).split('result:{policyname:')[1]
                col2 = cleanText(data[3]).split(':')[1]
                col3 = (cleanText(data[4]).split(':')[1]).split('}')[0]
                col4 = int(col2) - int(col3)
                col4 = str(col4)
            else:

                col1 = cleanText(data[3]).split('result:{policyname:')[1]
                col2 = cleanText(data[4]).split(':')[1]
                col3 = (cleanText(data[5]).split(':')[1]).split('}')[0]
                col4 = int(col2) - int(col3)
                col4 = str(col4)
            #################################################################
            # UBA 로그를 파싱하여 본문 데이터 작성 : 1. 메일스크린 정책 적용 현황

            HTML = HTML + """
<tr> 
<td style="border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 30px;" rowspan="1"> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;">""" + col1 + """</p> </td> 
<td style="border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 30px;" rowspan="1"> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;">""" + col2 + """</p> </td> 
<td style="border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 30px;" rowspan="1"> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;">""" + col3 + """</p> </td> 
<td style="border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 30px;" rowspan="1"> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;">""" + col4 + """</p> </td> 
</tr> """
            # 표 행 추가

#################################################################
# 테이블 닫기 : 1. 메일스크린 정책 적용 현황
HTML = HTML + """
</tbody> 
</table>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><br /></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><!-- //Table --><span style="font-family: 바탕;">&nbsp;</span></p>"""

#################################################################
# 제목 및 테이블 상단 작성 : 2. 부서별 발송 TOP 5
HTML = HTML + """
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong><span style="font-family: 바탕; font-size: 12pt;">2. 부서별 기밀정보 발송 TOP 5</span></strong></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 12pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;">&nbsp; ;&nbsp;기밀키워드를 가장 많이 사용한 부서 TOP5&nbsp;입니다. &nbsp;</span></span></span></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><br /></p>
<table style="border: 0px solid rgb(0, 0, 0); border-image: none; width: 305px; height: 185px; font-size: 10pt; border-collapse: collapse; background-color: rgb(255, 255, 255); resize: none;" border="1" cellspacing="0" cellpadding="0"> 
<tbody> 

<tr> 
<td style='border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 28px; font-family: ""; font-size: 10pt; background-color: rgb(204, 204, 204);'> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong>부서명</strong></p> </td> 
<td style='border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 28px; font-family: ""; font-size: 10pt; background-color: rgb(204, 204, 204);'> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong><span style='font-family: "바탕"; font-size: 10pt;'><strong>금주 발송 수량</strong></span></strong></p> </td> 
</tr> """

#################################################################
# 부서 TOP 5 수량 작성

#################################################################
# 본문작성을 위한 UBA 로그 연동 : 2. 부서별 발송 TOP 5

splunksql = """search index=mailscreen_group earliest="-7d@d" latest="@d" policyname = "올리브영*결재*"
 |rename sender_email as email_addr
 |fields sendtime, email_addr, header_subject, policyname
 |join email_addr
 [search index=ehr sourcetype=user_master comp_nm="*올리브영*" earliest="-1d@d" latest="@d" 
 |fields dept_nm, email_addr]
 |dedup sendtime, dept_nm, email_addr, header_subject, policyname
 |streamstats count as "RowCount1"
 |stats dc(RowCount1) as sendcnt by dept_nm
 |sort - sendcnt
 |head 5"""

PARAMS = {'search': splunksql,
          'earliest_time': earliest_time,
          'latest_time': latest_time,
          'output_mode': 'json'}
# 응답 요청
r = requests.post(url=logURL, headers=headerInfo, data=PARAMS, verify=False)

with open(now.strftime('%Y-%m-%d') + '_oyscreendept.csv', 'wb') as file:
    file.write(r.content)

with open(now.strftime('%Y-%m-%d') + '_oyscreendept.csv', "rt", encoding="UTF8") as infile:
    for line in infile:

        # 데이터를 쉼표 로 구분
        data = line.split(',\"')

        # preview:true는 버림. 왜 조회 결과가 2쌍으로 나오는지 모르곘음...
        # print(line)
        if cleanText(data[0]) == "{preview:false":
            # 조회 결과가 1개일 때, {"preview":false,"offset":0,"lastrow":true,"result":{"comp_nm":......
            if cleanText(data[2]) != "lastrow:true":
                # print(line)
                col1 = cleanText(data[2]).split('result:{dept_nm:')[1]
                col2 = (cleanText(data[3]).split(':')[1]).split('}')[0]

            else:

                col1 = cleanText(data[3]).split('result:{dept_nm:')[1]
                col2 = (cleanText(data[4]).split(':')[1]).split('}')[0]

            # 표 행 추가
            #################################################################
            # UBA 로그를 파싱하여 본문 데이터 작성 : 2. 부서별 발송 TOP 5

            HTML = HTML + """
<tr> 
<td style='border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 30px; font-family: ""; font-size: 10pt;' rowspan="1"> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;">""" + col1 + """</p> </td> 
<td style='border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 30px; font-family: ""; font-size: 10pt;' rowspan="1"> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;">""" + col2 + """</p> </td> 
</tr> """

#################################################################
# 테이블 닫기 : 2. 부서별 발송 TOP 5
HTML = HTML + """
</tbody> 
</table>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕;"><br /></span></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><br /></p>"""

#################################################################
# 제목 및 테이블 상단 작성
HTML = HTML + """
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong><span style="font-family: 바탕; font-size: 12pt;">3. 기밀정보별 사용 TOP 5</span></strong></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;">&nbsp; ; 가장 많이 사용된 기밀키워드 TOP5 입니다. </span></span></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;">&nbsp; ; 상세 내역은 첨부파일 내&nbsp;'3_' 시트에 별첨되었으며,&nbsp;</span></span></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;">&nbsp;&nbsp;&nbsp; 수신인 정보가 포함되어 아래의 표보다 많은 데이터가 포함되었습니다.&nbsp; </span></span></span></span></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;">&nbsp;&nbsp;&nbsp; (첨부파일에서 수신인정보(받는주소) 필드를 삭제한 후, 중복 제거 시 표와 동일한 수치 확인 가능)</span></span></span></span></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><br /></p>
<table style="border: 0px solid rgb(0, 0, 0); border-image: none; width: 305px; height: 185px; font-size: 10pt; border-collapse: collapse; background-color: rgb(255, 255, 255); resize: none;" border="1" cellspacing="0" cellpadding="0"> 
<tbody> 

<tr> 
<td style='border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 28px; font-family: ""; font-size: 10pt; background-color: rgb(204, 204, 204);'> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong>키워드</strong></p> </td> 
<td style='border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 28px; font-family: ""; font-size: 10pt; background-color: rgb(204, 204, 204);'> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong><span style='font-family: "바탕"; font-size: 10pt;'><strong>금주 사용 횟수</strong></span></strong></p> </td> 
</tr> """

#################################################################
# 키워드 사용 빈도 TOP 5

#################################################################
# 키워드 사용첨부파일 작성
splunksql = """search index=mailscreen_group     earliest="-7d@d" latest="@d" policyname = "올리브영*결재*"
|eval header_subject = replace(header_subject, ",", ".")
|eval keyword_pre = replace(keyword_pre, ",", ".")
|eval keyword_ptn = replace(keyword_ptn, ",", ".")
|eval keyword_post = replace(keyword_post, ",", ".")
|fields sendtime, sender_email, receiver_email, header_subject, policyname, keyword_pre, keyword_ptn, keyword_post
|join keyword_ptn
[search index=mailscreen_group policyname = "올리브영*결재*" earliest="-7d@d" latest="@d"
 |fields sendtime, sender_email, header_subject, policyname, keyword_pre, keyword_ptn, keyword_post
 |dedup sendtime, sender_email, header_subject, policyname, keyword_pre, keyword_ptn, keyword_post
 |streamstats count as "RowCount"
 |stats dc(RowCount) as sndcnt by keyword_ptn
 |sort - sndcnt
 |head 5]
|join sender_email
[search index=ehr sourcetype=user_master comp_nm="*올리브영*" earliest="-1d@d" latest="@d" 
 |rename email_addr as sender_email
 |fields dept_nm, sender_email, user_nm]
|dedup sendtime, dept_nm, user_nm, sender_email, receiver_email, header_subject, policyname, keyword_pre, keyword_ptn, keyword_post
|table sendtime, dept_nm, user_nm, sender_email, receiver_email, header_subject, policyname, keyword_pre, keyword_ptn, keyword_post
|sort - sndcnt, user_nm, sendtime"""

PARAMS = {'search': splunksql,
          'earliest_time': earliest_time,
          'latest_time': latest_time,
          'output_mode': 'json'}
# 응답 요청
r = requests.post(url=logURL, headers=headerInfo, data=PARAMS, verify=False)

# UBA 로그 다운로드
with open(now.strftime('%Y-%m-%d') + '_oyscreenkeyuserdetail.csv', 'wb') as file:
    file.write(r.content)

# xlsx 생성 시작
wb = Workbook()
# 시트 생성
sheet1 = wb.active
sheet1.title = '3_기밀정보별사용TOP5'
sheet2 = wb.create_sheet('4_임직원별기밀정보사용TOP5')

# 컬럼 필드 너비 조정
for column in ascii_uppercase:
    if (column == 'A'):
        sheet1.column_dimensions[column].width = 18
        sheet2.column_dimensions[column].width = 18
    elif (column == 'B'):
        sheet1.column_dimensions[column].width = 16
        sheet2.column_dimensions[column].width = 16
    elif (column == 'C'):
        sheet1.column_dimensions[column].width = 9
        sheet2.column_dimensions[column].width = 9
    elif (column == 'D'):
        sheet1.column_dimensions[column].width = 17
        sheet2.column_dimensions[column].width = 17
    elif (column == 'E'):
        sheet1.column_dimensions[column].width = 17
        sheet2.column_dimensions[column].width = 17
    elif (column == 'F'):
        sheet1.column_dimensions[column].width = 45
        sheet2.column_dimensions[column].width = 45
    elif (column == 'G'):
        sheet1.column_dimensions[column].width = 18
        sheet2.column_dimensions[column].width = 18
    elif (column == 'H'):
        sheet1.column_dimensions[column].width = 33
        sheet2.column_dimensions[column].width = 33
    elif (column == 'I'):
        sheet1.column_dimensions[column].width = 10
        sheet2.column_dimensions[column].width = 10
    elif (column == 'J'):
        sheet1.column_dimensions[column].width = 33
        sheet2.column_dimensions[column].width = 33
    else:
        sheet1.column_dimensions[column].width = 10
        sheet2.column_dimensions[column].width = 10

# 시트1
title = ["발송시각", "부서명", "임직원명", "보내는주소", "받는주소", "메일제목", "적용정책명", "키워드앞", "키워드", "키워드뒤"]
writeColName(title, 1, sheet1)

# UBA 로그 읽기
with open(now.strftime('%Y-%m-%d') + '_oyscreenkeyuserdetail.csv', "rt", encoding="UTF8") as infile:
    row_num = 2
    for line in infile:

        # 데이터를 쉼표 로 구분
        data = line.split(',\"')

        # preview:true는 버림. 왜 조회 결과가 2쌍으로 나오는지 모르곘음...
        # print(line)
        if cleanText(data[0]) == "{preview:false":
            # 조회 결과가 1개일 때, {"preview":false,"offset":0,"lastrow":true,"result":{"comp_nm":......
            if cleanText(data[2]) != "lastrow:true":

                col1 = cleanText(data[2]).split('result:{sendtime:')[1]
                col2 = cleanText(data[3]).split('dept_nm:')[1]
                col3 = cleanText(data[4]).split('user_nm:')[1]
                col4 = cleanText(data[5]).split('sender_email:')[1]
                col5 = cleanText(data[6]).split('receiver_email:')[1]
                col6 = cleanText(data[7]).split('header_subject:')[1]
                col7 = cleanText(data[8]).split('policyname:')[1]
                col8 = cleanText(data[9]).split('keyword_pre:')[1]
                col9 = cleanText(data[10]).split('keyword_ptn:')[1]
                col10 = (cleanText(data[11]).split('keyword_post:')[1]).split('}')[0]
            else:

                col1 = cleanText(data[3]).split('result:{sendtime:')[1]
                col2 = cleanText(data[4]).split('dept_nm:')[1]
                col3 = cleanText(data[5]).split('user_nm:')[1]
                col4 = cleanText(data[6]).split('sender_email:')[1]
                col5 = cleanText(data[7]).split('receiver_email:')[1]
                col6 = cleanText(data[8]).split('header_subject:')[1]
                col7 = cleanText(data[9]).split('policyname:')[1]
                col8 = cleanText(data[10]).split('keyword_pre:')[1]
                col9 = cleanText(data[11]).split('keyword_ptn:')[1]
                col10 = (cleanText(data[12]).split('keyword_post:')[1]).split('}')[0]

            contetnt = [col1, col2, col3, col4, col5, col6, col7, col8, col9, col10]
            saveContent(contetnt, sheet1)
            row_num += 1

        #################################################################
#################################################################
# 본문작성을 위한 UBA 로그 연동
splunksql = """search index=mailscreen_group policyname = "올리브영*결재*" earliest="-7d@d" latest="@d"
 |fields sendtime, sender_email, header_subject, policyname, keyword_pre, keyword_ptn, keyword_post
 |dedup sendtime, sender_email, header_subject, policyname, keyword_pre, keyword_ptn, keyword_post
 |streamstats count as "RowCount"
 |stats dc(RowCount) as sndcnt by keyword_ptn
 |sort - sndcnt
 |head 5"""

PARAMS = {'search': splunksql,
          'earliest_time': earliest_time,
          'latest_time': latest_time,
          'output_mode': 'json'}
# 응답 요청
r = requests.post(url=logURL, headers=headerInfo, data=PARAMS, verify=False)

# UBA 로그 다운로드
with open(now.strftime('%Y-%m-%d') + '_oyscreenkey.csv', 'wb') as file:
    file.write(r.content)

with open(now.strftime('%Y-%m-%d') + '_oyscreenkey.csv', "rt", encoding="UTF8") as infile:
    for line in infile:

        # 데이터를 쉼표 로 구분
        data = line.split(',\"')

        # preview:true는 버림. 왜 조회 결과가 2쌍으로 나오는지 모르곘음...
        # print(line)
        if cleanText(data[0]) == "{preview:false":
            # 조회 결과가 1개일 때, {"preview":false,"offset":0,"lastrow":true,"result":{"comp_nm":......
            if cleanText(data[2]) != "lastrow:true":
                # print(line)
                col1 = cleanText(data[2]).split('result:{keyword_ptn:')[1]
                col2 = (cleanText(data[3]).split(':')[1]).split('}')[0]

            else:

                col1 = cleanText(data[3]).split('result:{keyword_ptn:')[1]
                col2 = (cleanText(data[4]).split(':')[1]).split('}')[0]

            # 표 행 추가
            HTML = HTML + """
<tr> 
<td style='border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 30px; font-family: ""; font-size: 10pt;' rowspan="1"> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;">""" + col1 + """</p> </td> 
<td style='border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 30px; font-family: ""; font-size: 10pt;' rowspan="1"> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;">""" + col2 + """</p> </td> 
</tr> """

HTML = HTML + """
</tbody> 
</table>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><em><br /></em></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong><span style="font-size: 12pt;"><br /></span></strong></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong><span style="font-family: 바탕; font-size: 12pt;">4. 임직원별 기밀정보 사용 TOP 5</span></strong></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;">&nbsp; ; 기밀키워드를 가장 많이 사용한 임직원&nbsp;TOP5 입니다. </span></span></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;">&nbsp; ; 상세 내역은 첨부파일 내&nbsp;'4_' 시트에 별첨되었으며,&nbsp;</span></span></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;">&nbsp;&nbsp;&nbsp; 수신인 정보가 포함되어 아래의 표보다 많은 데이터가 포함되었습니다.&nbsp; </span></span></span></span></p>

<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;">&nbsp;&nbsp;&nbsp; (첨부파일에서 수신인정보(받는주소) 필드를 삭제한 후, 중복 제거 시 표와 동일한 수치 확인 가능)</span></span></span></span></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕;">&nbsp;</span></p>
<table style="border: 0px solid rgb(0, 0, 0); border-image: none; width: 457px; height: 185px; font-size: 10pt; border-collapse: collapse; background-color: rgb(255, 255, 255); resize: none;" border="1" cellspacing="0" cellpadding="0"> 
<tbody> 

<tr> 
<td style='border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 28px; font-family: ""; font-size: 10pt; background-color: rgb(204, 204, 204);'> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong>부서명</strong></p> </td> 
<td style='border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 28px; font-family: ""; font-size: 10pt; background-color: rgb(204, 204, 204);'> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;"><strong>임직원</strong></p> </td> 
<td style='border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 28px; font-family: ""; font-size: 10pt; background-color: rgb(204, 204, 204);'> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style='font-family: "바탕"; font-size: 10pt;'><strong>금주 사용 횟수</strong></span></p> </td> 
</tr> """

#################################################################
# 임직원별 키워드 사용 TOP 5

#################################################################
# 임직원별 키워드 사용첨부파일 작성
splunksql = """search index=mailscreen_group     earliest="-7d@d" latest="@d" policyname = "올리브영*결재*"
|eval header_subject = replace(header_subject, ",", ".")
|eval keyword_pre = replace(keyword_pre, ",", ".")
|eval keyword_ptn = replace(keyword_ptn, ",", ".")
|eval keyword_post = replace(keyword_post, ",", ".")
|fields sendtime, sender_email, receiver_email, header_subject, policyname, keyword_pre, keyword_ptn, keyword_post
|join sender_email
[search index=mailscreen_group     earliest="-7d@d" latest="@d" policyname = "올리브영*결재*"
 |eval header_subject = replace(header_subject, ",", ".")
 |eval keyword_pre = replace(keyword_pre, ",", ".")
 |eval keyword_ptn = replace(keyword_ptn, ",", ".")
 |eval keyword_post = replace(keyword_post, ",", ".")
 |fields sendtime, sender_email, header_subject, policyname, keyword_pre, keyword_ptn, keyword_post
 |dedup sendtime, sender_email, header_subject, policyname, keyword_pre, keyword_ptn, keyword_post
 |streamstats count as "RowCount"
 |stats dc(RowCount) as sndcnt by sender_email
 |sort - sndcnt
 |head 5]
|join sender_email
[search index=ehr sourcetype=user_master comp_nm="*올리브영*" earliest="-1d@d" latest="@d" 
 |rename email_addr as sender_email
 |fields dept_nm, sender_email, user_nm]
|dedup sendtime, dept_nm, user_nm, sender_email, receiver_email, header_subject, policyname, keyword_pre, keyword_ptn, keyword_post
|table sendtime, dept_nm, user_nm, sender_email, receiver_email, header_subject, policyname, keyword_pre, keyword_ptn, keyword_post
|sort - sndcnt, user_nm, sendtime"""

PARAMS = {'search': splunksql,
          'earliest_time': earliest_time,
          'latest_time': latest_time,
          'output_mode': 'json'}
# 응답 요청
r = requests.post(url=logURL, headers=headerInfo, data=PARAMS, verify=False)

# UBA 로그 다운로드
with open(now.strftime('%Y-%m-%d') + '_oyscreenkeyuserdetail.csv', 'wb') as file:
    file.write(r.content)

title = ["발송시각", "부서명", "임직원명", "보내는주소", "받는주소", "메일제목", "적용정책명", "키워드앞", "키워드", "키워드뒤"]
writeColName(title, 1, sheet2)

# UBA 로그 읽기
with open(now.strftime('%Y-%m-%d') + '_oyscreenkeyuserdetail.csv', "rt", encoding="UTF8") as infile:
    row_num = 2
    for line in infile:

        # 데이터를 쉼표 로 구분
        data = line.split(',\"')

        # preview:true는 버림. 왜 조회 결과가 2쌍으로 나오는지 모르곘음...
        # print(line)
        if cleanText(data[0]) == "{preview:false":
            # 조회 결과가 1개일 때, {"preview":false,"offset":0,"lastrow":true,"result":{"comp_nm":......
            if cleanText(data[2]) != "lastrow:true":

                col1 = cleanText(data[2]).split('result:{sendtime:')[1]
                col2 = cleanText(data[3]).split('dept_nm:')[1]
                col3 = cleanText(data[4]).split('user_nm:')[1]
                col4 = cleanText(data[5]).split('sender_email:')[1]
                col5 = cleanText(data[6]).split('receiver_email:')[1]
                col6 = cleanText(data[7]).split('header_subject:')[1]
                col7 = cleanText(data[8]).split('policyname:')[1]
                col8 = cleanText(data[9]).split('keyword_pre:')[1]
                col9 = cleanText(data[10]).split('keyword_ptn:')[1]
                col10 = (cleanText(data[11]).split('keyword_post:')[1]).split('}')[0]
            else:

                col1 = cleanText(data[3]).split('result:{sendtime:')[1]
                col2 = cleanText(data[4]).split('dept_nm:')[1]
                col3 = cleanText(data[5]).split('user_nm:')[1]
                col4 = cleanText(data[6]).split('sender_email:')[1]
                col5 = cleanText(data[7]).split('receiver_email:')[1]
                col6 = cleanText(data[8]).split('header_subject:')[1]
                col7 = cleanText(data[9]).split('policyname:')[1]
                col8 = cleanText(data[10]).split('keyword_pre:')[1]
                col9 = cleanText(data[11]).split('keyword_ptn:')[1]
                col10 = (cleanText(data[12]).split('keyword_post:')[1]).split('}')[0]

            contetnt = [col1, col2, col3, col4, col5, col6, col7, col8, col9, col10]
            # print(row_num)
            saveContent(contetnt, sheet2)
            row_num += 1

# oyf.close()
wb.save(now.strftime('%Y-%m-%d') + '_올리브영_기밀정보사용상세.xlsx')
#################################################################

splunksql = """search index=mailscreen_group     earliest="-7d@d" latest="@d" policyname = "올리브영*결재*"
 |fields sendtime, sender_email, header_subject, policyname, keyword_pre, keyword_ptn, keyword_post
 |join sender_email
 [search index=ehr sourcetype=user_master comp_nm="*올리브영*" earliest="-1d@d" latest="@d" 
 |rename email_addr as sender_email 
 |fields dept_nm, email_addr, user_nm, sender_email]
 |dedup sendtime, dept_nm, user_nm, sender_email, header_subject, policyname, keyword_pre, keyword_ptn, keyword_post
 |streamstats count as "RowCount"
 |stats values(dept_nm), values(user_nm), dc(RowCount) as sndcnt by sender_email
 |sort - sndcnt, values(user_nm)
 |head 5"""

PARAMS = {'search': splunksql,
          'earliest_time': earliest_time,
          'latest_time': latest_time,
          'output_mode': 'json'}
# 응답 요청
r = requests.post(url=logURL, headers=headerInfo, data=PARAMS, verify=False)

# UBA 로그 다운로드
with open(now.strftime('%Y-%m-%d') + '_oyscreenuserpt.csv', 'wb') as file:
    file.write(r.content)

# UBA 로그 읽기
with open(now.strftime('%Y-%m-%d') + '_oyscreenuserpt.csv', "rt", encoding="UTF8") as infile:
    for line in infile:

        # 데이터를 쉼표 로 구분
        data = line.split(',\"')

        # preview:true는 버림. 왜 조회 결과가 2쌍으로 나오는지 모르곘음...
        # print(line)
        if cleanText(data[0]) == "{preview:false":
            # 조회 결과가 1개일 때, {"preview":false,"offset":0,"lastrow":true,"result":{"comp_nm":......
            if cleanText(data[2]) != "lastrow:true":
                # print(line)
                col1 = cleanText(data[3]).split(':')[1]
                col2 = cleanText(data[4]).split(':')[1]
                col3 = (cleanText(data[5]).split(':')[1]).split('}')[0]

            else:

                col1 = cleanText(data[4]).split(':')[1]
                col2 = cleanText(data[5]).split(':')[1]
                col3 = (cleanText(data[6]).split(':')[1]).split('}')[0]

            # 표 행 추가
            HTML = HTML + """
<tr> 
<td style='border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 30px; font-family: ""; font-size: 10pt;' rowspan="1"> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;">""" + col2 + """</p> </td> 
<td style='border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 30px; font-family: ""; font-size: 10pt;' rowspan="1"> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;">""" + col1 + """</p> </td> 
<td style='border: 1px solid rgb(0, 0, 0); border-image: none; width: 151px; height: 30px; font-family: ""; font-size: 10pt;' rowspan="1"> 
<p class="ce_exstyle" style="text-align: center; line-height: 1.2; font-family: 바탕; font-size: 10pt;">""" + col3 + """</p> </td> 
</tr> """

HTML = HTML + """
</tbody> 
</table>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-size: 10pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;"><br /></span></span></span></span></span></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-size: 10pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 11pt;">&nbsp;</span></span></span></span></span></p>
<p class="ce_exstyle" style="line-height: 1.2; font-family: 바탕; font-size: 10pt;"><span style="font-family: 바탕; font-size: 10pt;"><em><span style="font-size: 10pt;"><span style="font-style: normal;">※ 본 메일은 발신전용으로 메일 수신이 불가능합니다. 문의사항은 "플랫폼보안" DL로 문의주시기 바랍니다. </span></span></em></span></p>
</body>
</html>
"""
# print(HTML)

SERVER = 'as2.cj.net:25'
# RECEIVER_EMAIL = ['minsook.kim1@cj.net', 'mskim2018@cj.net']
SENDER_EMAIL = 'platform@cjautoreport.com'

with open('recipients.txt', 'r') as f:
    RECEIVER_EMAIL = []
    for line in f:
        RECEIVER_EMAIL.append(line)

    # message = MIMEMultipart("alternative", None, [MIMEText(HTML, 'html')])
# message['Subject'] = '[메일스크린] ' + latest_time + '시 변조 현황 ' + str(checkc) + '건, 변조 의심 현황 ' + str(checkc2) + '건 입니다.'
# message['From'] = SENDER_EMAIL
# message['To'] = ", ".join(RECEIVER_EMAIL)

# 정의
# def _generate_message() -> MIMEMultipart:

# message = MIMEMultipart("alternative", None, [MIMEText(HTML, 'html')])
message = MIMEBase('multipart', 'mixed')
# 본문 삽입
message.attach(MIMEText(HTML, 'html'))

############### ↓ 첨부파일이 없다면 삭제 가능  ↓ ########################
# 첨부파일 경로/이름 지정하기
attachfile = now.strftime('%Y-%m-%d') + '_올리브영_기밀정보사용상세.xlsx'
attachment = open(attachfile, 'rb')

part = MIMEBase('application', 'octet-stream')
part.set_payload((attachment).read())
encoders.encode_base64(part)
part.add_header('Content-Disposition', 'attachment', filename=attachfile)
message.attach(part)

# print(os.path.basename(filename))
############### ↑ 첨부파일이 없다면 삭제 가능  ↑ ########################

message['Subject'] = '[주간리포트] 메일스크린 운영 현황 (' + oneweekday + ' ~ ' + yesterday + ')'
message['From'] = SENDER_EMAIL
message['To'] = ", ".join(RECEIVER_EMAIL)
# return message

# message = _generate_message()
server = smtplib.SMTP(SERVER)
server.ehlo()
server.starttls()
# server.login(SENDER_EMAIL, SENDER_PASSWORD) 인증 필요 없음
server.sendmail(SENDER_EMAIL, RECEIVER_EMAIL, message.as_string())
server.quit()

r.connection.close()

