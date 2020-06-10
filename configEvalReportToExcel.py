
import boto3
import json
import datetime
import requests
import time
import win32com.client
import os
from pywintypes import com_error
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Color, NamedStyle
from openpyxl.chart import BarChart, Reference
from openpyxl.styles.borders import Border, Side
#pip install pywin32

#create config client
client = boto3.client('config')

# Create Work Book(Whole Excel file)
wb = Workbook()
ws = wb.active
ws.title = 'init'
# data save from
column_num = 3
#compliant summary시 사용될 변수
comp_num = 0
noncomp_num = 0
excelData = []

paintstyle = NamedStyle(name="border")
bd = Side(border_style='thin', color="000000")
paintstyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)
paintstyle.alignment = Alignment(vertical='center')

# Write Column Name
def writeColName(cName, row, ws):
    column_char = 'a'
    #write column name
    for name in cName:
        ws[column_char + str(row)] = name
        ws[column_char + str(row)].style = paintstyle
        ws[column_char + str(row)].alignment = Alignment(horizontal='center', vertical='center')
        column_char = chr(ord(column_char) + 1)

# Save Compliance Data to Excel
def saveContent(cName, ws):
    column_char = 'a'
    global column_num
    for name in cName:
        ws[column_char + str(column_num)] = name
        ws[column_char + str(column_num)].style = paintstyle
        ws[column_char + str(column_num)].alignment = Alignment(vertical='center', wrap_text=True)
        column_char = chr(ord(column_char) + 1)
    column_num += 1

#시트 첫행 타이틀 굵게, 가운데정렬, 배경색설정
def setTitleCell(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(patternType='solid', fgColor=Color("FFFF00"))
    return cell

#Page one for AWS Config Rule description(규칙 항목 및 설명)
def sheetOne():
    ruledetailpaginator = client.get_paginator('describe_config_rules')
    response_iterator = ruledetailpaginator.paginate()
    global ws
    column_num = 3
    ws.merge_cells('A1:B1')
    ws['A1'] = "규칙 항목 및 설명"
    ws['A1'].style = paintstyle
    ws['B1'].style = paintstyle
    setTitleCell(ws.cell(row=1, column=1))
    ws.title = 'ConfigRuleList'

    # call export func by configrule for loop
    for configrule in response_iterator:
        for rulename in configrule['ConfigRules']:
            # init column num for each Config Rule Result
            excelData = []
            excelData.append(rulename['ConfigRuleName'])
            excelData.append(rulename['Description'])
            saveContent(excelData, ws)

    #put the data title on sheet1
    excelColName = ['AWS Config 규칙명', 'Rule Description']
    writeColName(excelColName,2, ws)
    ws.row_dimensions[1].height = 25
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

    print('first sheet done')

#Sheet 2 for Evaluated resources - 평가된 리소스 현황
def sheetTwo():
    ws = wb.create_sheet('Evaluated Resources')
    discovered_rsc = client.get_discovered_resource_counts()
    global column_num
    column_num = 4
    ws.merge_cells('A1:B1')
    ws['A1'] = "Evaluation Resources(평가 리소스 현황)"
    ws['A1'].style = paintstyle
    ws['B1'].style = paintstyle
    setTitleCell(ws.cell(row=1, column=1))
    ws['A2'].style = paintstyle
    ws['A2'].value = '총 리소스 수'
    ws['B2'].style = paintstyle
    ws['B2'].value = discovered_rsc['totalDiscoveredResources']

    for resource in discovered_rsc['resourceCounts']:
        excelData = []
        #print(resource)
        excelData.append(resource['resourceType'])
        excelData.append(resource['count'])
        saveContent(excelData, ws)

    excelColName = ['리소스 유형', '합계']
    writeColName(excelColName, 3, ws)
    ws.row_dimensions[1].height = 25
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15

    print('second sheet done')



#Sheet 3 for evaluation summary by config rule - Config Rule 별 평과 결과 막대그래프
def sheetThree():
    ws = wb.create_sheet('Evaluation Summary')
    ruleSummaryRes = client.get_compliance_summary_by_config_rule()
    global column_num
    column_num = 2
    ws.merge_cells('A1:C1')
    ws['A1'] = "Evaluation Summary(by config rule)"
    #border line



    comp = ruleSummaryRes['ComplianceSummary']['CompliantResourceCount']['CappedCount']
    noncomp = ruleSummaryRes['ComplianceSummary']['NonCompliantResourceCount']['CappedCount']
    rows = [
        ('', 'compliant', 'non_compliant'),
        ('Rule count',comp,noncomp)
    ]

    for row in rows:
        ws.append(row)

    bc = BarChart()
    bc.type = "col"
    bc.style = 10
    bc.title = "Evaluation Summary"
    bc.y_axis.title = 'Rule Count'
    bc.x_axis.title = 'Compliant Status'

    data = Reference(ws, min_col=2, min_row=1, max_row=4, max_col=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=4)
    bc.add_data(data, titles_from_data=True)
    bc.set_categories(cats)
    bc.shape = 4
    bc.width = 8
    bc.legend = None
    ws.add_chart(bc, "A6")

    rownum=0
    while rownum < 3:
        ws.cell(row=rownum+1, column=1).style = paintstyle
        ws.cell(row=rownum+1, column=2).style = paintstyle
        ws.cell(row=rownum+1, column=3).style = paintstyle
        rownum += 1
    setTitleCell(ws.cell(row=1, column=1))

    ws.row_dimensions[1].height = 25
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    print('third sheet done')

#sheet 4 for numbers of compliance result - 평가 항목 별 compliant 요약
def sheetFour():
    # get_compliance_details_by_config_rule paginator
    ruledetailpaginator = client.get_paginator('describe_config_rules')
    response_iterator = ruledetailpaginator.paginate()
    paginator = client.get_paginator('get_compliance_details_by_config_rule')
    # Worksheet 생성
    ws = wb.create_sheet('Summary(Number)')
    global column_num
    rulenametemp =''
    column_num = 3
    ws.merge_cells('A1:C1')
    ws['A1'] = "평가 항목 별 요약"
    #borderline
    ws['A1'].style = paintstyle
    ws['B1'].style = paintstyle
    ws['C1'].style = paintstyle
    setTitleCell(ws.cell(row=1, column=1))

    #describe rule의 rule name 추출해서 리소스들 get comp detail 정보 추출
    for configrule in response_iterator:
        for rulename in configrule['ConfigRules']:
            responseConfigPage = paginator.paginate(
                ConfigRuleName=rulename['ConfigRuleName']
            )
            # 룰별 객체
            for page in responseConfigPage:
                global excelData, comp_num, noncomp_num
                try:
                    if not rulenametemp :
                        rulenametemp = page['EvaluationResults'][0]['EvaluationResultIdentifier']['EvaluationResultQualifier']['ConfigRuleName']
                        excelData.append(rulenametemp)

                    for user in page['EvaluationResults']:
                        if user['ComplianceType'] == 'NON_COMPLIANT':
                            noncomp_num += 1
                        elif user['ComplianceType'] == 'COMPLIANT':
                            comp_num += 1
                    try:
                        #NextToken 값 존재(pagination)
                        if page['NextToken'] in page:
                            pass
                    except:
                        #룰네임존재, 리소스 pagination 아님 엑셀저장
                        #print(comp_num)
                        #print(noncomp_num)
                        excelData.append(comp_num)
                        excelData.append(noncomp_num)
                        saveContent(excelData, ws)
                        comp_num = 0
                        noncomp_num = 0
                        rulenametemp = ''
                        excelData = []

                except:
                    if(rulenametemp == '') :
                        continue
                    else :
                        #마지막 페이지, nexttoken 없음 엑셀저장
                        #print(comp_num)
                        #print(noncomp_num)
                        excelData.append(comp_num)
                        excelData.append(noncomp_num)
                        saveContent(excelData, ws)
                        comp_num = 0
                        noncomp_num = 0
                        rulenametemp =''
                        excelData = []

    #put the data title on sheet1
    excelColName = ['AWS Config 규칙명', 'COMPLIANT', 'NON COMPLIANT']
    writeColName(excelColName,2, ws)
    ws.row_dimensions[1].height = 25
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws['A'+str(ws.max_row+2)] = '평가되는 리소스가 없는 룰은 제외됩니다.'

    print('last sheet done')


###########################################Main############################################

print('creating excel file..')
sheetOne()
sheetTwo()
sheetThree()
sheetFour()
time.sleep(2)
wb.save("AWS_Config_Report.xlsx")
print('Excel file saved')

WB_PATH = 'AWS_Config_Report.xlsx'
PATH_TO_PDF = 'AWS_Config_Report.pdf'

excel = win32com.client.Dispatch("Excel.application")

try:
    print('Starting convert to PDF')
    wb=excel.Workbooks.Open(os.path.abspath('AWS_Config_Report.xlsx'))
    time.sleep(1)
    ws_list = ['ConfigRuleList', 'Evaluated Resources','Evaluation Summary','Summary(Number)']
    wb.WorkSheets(ws_list).Select()
    excel.ActiveSheet.ExportAsFixedFormat(0, os.path.abspath(PATH_TO_PDF))

except com_error as e:
    print('converting failed.')
    print(e)
else:
    print('PDF file saved')
finally:
    wb.Close(False)
    excel.Quit()

print("Automatically closing...")
###########################################################################################