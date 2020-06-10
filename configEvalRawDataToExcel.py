
import boto3
import json
import datetime
import requests
from openpyxl import Workbook
from openpyxl import load_workbook

#create config client
client = boto3.client('config')

# Create Work Book(Whole Excel file)
wb = Workbook()
ws = wb.active
ws.title = 'init'
# data save from
column_num = 2

# Write Column Name
def writeColName(cName):
    column_char = 'a'
    for name in cName:
        ws[column_char + '1'] = name
        column_char = chr(ord(column_char) + 1)

# Save Compliance Data to Excel
def saveContent(cName):
    column_char = 'a'
    global column_num
    for name in cName:
        ws[column_char + str(column_num)] = name
        column_char = chr(ord(column_char) + 1)
    column_num += 1

#export AWS Config evaluation result to excel
def exportExcelByCfgRuleName(configrulename, desc):
    global ws

    if(ws.title == 'init') :
        ws.title = 'Config Result'
    else :
        ws = wb.create_sheet('Config Result')

    #get detail paginator
    paginator = client.get_paginator('get_compliance_details_by_config_rule')
    responseConfigPage = paginator.paginate(
        ConfigRuleName=configrulename
    )

    for page in responseConfigPage:
        for user in page['EvaluationResults']:
            excelData = []
            excelData.append(user['EvaluationResultIdentifier']['EvaluationResultQualifier']['ConfigRuleName'])
            excelData.append(user['EvaluationResultIdentifier']['EvaluationResultQualifier']['ResourceType'])
            excelData.append(user['EvaluationResultIdentifier']['OrderingTimestamp'])
            excelData.append(user['ResultRecordedTime'])
            excelData.append(usernameDic.get(user['EvaluationResultIdentifier']['EvaluationResultQualifier']['ResourceId']))
            excelData.append(user['EvaluationResultIdentifier']['EvaluationResultQualifier']['ResourceId'])
            excelData.append(user['ComplianceType'])
            if(user['ComplianceType']=='NON_COMPLIANT'):
                excelData.append(desc)
            saveContent(excelData)

    excelColName = ['AWS Config 규칙명', '리소스 타입', 'Time of the event', 'Result Recorded Time','UserName','ResourceId','컴플라이언스 결과', 'Rule Description']
    writeColName(excelColName)

#################################Main############################################
#call config rules
rulepaginator = client.get_paginator('describe_config_rules')
response_iterator = rulepaginator.paginate()
userinfo = client.list_discovered_resources(
    resourceType='AWS::IAM::User'
)

#user name join from list_discovered_resources
usernameDic = {}
for user_info in userinfo['resourceIdentifiers']:
    usernameDic[user_info['resourceId']] = user_info['resourceName']

#call export func by configrule for loop
for configrule in response_iterator:
    for rulename in configrule['ConfigRules']:
        #init column num for each Config Rule Result
        column_num = 2
        exportExcelByCfgRuleName(rulename['ConfigRuleName'], rulename['Description'])

wb.save("AWS_Config_Raw_Data.xlsx")
print("done")
