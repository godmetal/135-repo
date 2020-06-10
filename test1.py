
import boto3
import json
import datetime
import requests
from openpyxl import Workbook
from openpyxl import load_workbook

#create config client
client = boto3.client('config')

# Create Work Book(Whole Excel file)
wb = Workbook()
ws = wb.active
ws.title = 'init'
# data save from
column_num = 2

# Write Column Name
def writeColName(cName):
    column_char = 'a'
    for name in cName:
        ws[column_char + '1'] = name
        column_char = chr(ord(column_char) + 1)

# Save Compliance Data to Excel
def saveContent(cName):
    column_char = 'a'
    global column_num
    for name in cName:
        ws[column_char + str(column_num)] = name
        column_char = chr(ord(column_char) + 1)
    column_num += 1

#export AWS Config Rule result to excel
def exportExcelByCfgRuleName(configrulename, desc):
    global ws

    if(ws.title == 'init') :
        ws.title = 'Config 결과'
    else :
        ws = wb.create_sheet('Config 결과')

    #get detail paginator
    paginator = client.get_paginator('get_compliance_details_by_config_rule')
    responseConfigPage = paginator.paginate(
        ConfigRuleName=configrulename
    )
    print(responseConfigPage)
    for page in responseConfigPage:
        print(page)
        for user in page['EvaluationResults']:
            excelData = []
            excelData.append(user['EvaluationResultIdentifier']['EvaluationResultQualifier']['ConfigRuleName'])
            excelData.append(user['EvaluationResultIdentifier']['EvaluationResultQualifier']['ResourceType'])
            excelData.append(user['EvaluationResultIdentifier']['OrderingTimestamp'])
            excelData.append(user['ResultRecordedTime'])
            excelData.append(usernameDic.get(user['EvaluationResultIdentifier']['EvaluationResultQualifier']['ResourceId']))
            excelData.append(user['EvaluationResultIdentifier']['EvaluationResultQualifier']['ResourceId'])
            excelData.append(user['ComplianceType'])
            if(user['ComplianceType']=='NON_COMPLIANT'):
                excelData.append(desc)
            saveContent(excelData)

    excelColName = ['AWS Config 규칙명', '리소스 타입', 'Time of the event', 'Result Recorded Time','UserName','ResourceId','컴플라이언스 결과', 'Rule Description']
    writeColName(excelColName)


#################################Main############################################
#call config rules
response = client.describe_config_rules()
userinfo = client.list_discovered_resources(
    resourceType='AWS::IAM::User'
)

#user name join from list_discovered_resources
usernameDic = {}
for user_info in userinfo['resourceIdentifiers']:
    usernameDic[user_info['resourceId']] = user_info['resourceName']

#just print response
print(response)

#call slack func by configrule for loop
for configrule in response['ConfigRules']:
    column_num = 2
    exportExcelByCfgRuleName(configrule['ConfigRuleName'], configrule['Description'])

wb.save("configtest.xlsx")
